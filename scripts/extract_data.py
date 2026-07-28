#!/usr/bin/env python3
"""
Extract the Primelaze Unified Dashboard and Esthemax price workbooks into a
single clean JSON payload consumed by the web app.

Usage:
    python scripts/extract_data.py

Outputs:
    src/data/data.json      -- pretty JSON (source of truth, easy to diff)
    assets/js/data.js       -- `window.APP_DATA = {...}` (loaded by the app,
                               works over file:// with no fetch/CORS issues)

The parsers below are intentionally targeted at the known layout of each sheet
rather than being fully generic — the workbooks are hand-built dashboards with
section headers and merged cells, so a generic table reader produces noise.
"""
import json
import os
import datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
WB_DIR = os.path.join(ROOT, "source_workbooks")
DASHBOARD = os.path.join(WB_DIR, "Primelaze_Unified_Dashboard_FY2627.xlsx")
PRICES = os.path.join(WB_DIR, "Esthemax_prices_working.xlsx")
ORDER = os.path.join(WB_DIR, "Esthemax_Order_Calculation.xlsx")
DEMO = os.path.join(WB_DIR, "Demo_Machine_Booking_Status.xlsx")
CHALLAN = os.path.join(WB_DIR, "Delivery_Challan.xlsx")


def rows_of(ws):
    """Return list of row-lists (1-indexed friendly: rows_of(ws)[i] == row i+1)."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def clean(v):
    if isinstance(v, str):
        return v.strip()
    return v


def num(v):
    """Round floats a touch to kill spreadsheet noise, keep ints as ints."""
    if isinstance(v, float):
        if v == int(v):
            return int(v)
        return round(v, 2)
    return v


# --------------------------------------------------------------------------
# Dashboard workbook
# --------------------------------------------------------------------------
def parse_kpis(ws):
    r = rows_of(ws)
    # R7 (index 6): [None, 229, None, 120, None, 12, None, 4]
    vals = r[6]
    labels = r[7]
    return {
        "devices": num(vals[1]),
        "celluma": num(vals[3]),
        "activeReps": num(vals[5]),
        "vacant": num(vals[7]),
        "devicesNote": clean(labels[1]),
        "cellumaNote": clean(labels[3]),
        "repsNote": clean(labels[5]),
        "vacantNote": clean(labels[7]),
    }


def parse_master_targets(ws):
    """Rows 14..46: zone headers, HQ headers, and numbered salesperson rows."""
    r = rows_of(ws)
    zones = []
    cur_zone = None
    cur_hq = None
    for row in r[13:47]:
        c1 = clean(row[1])
        if c1 is None:
            continue
        text = str(c1)
        if text.startswith("◆") or "ZONE" in text and text.startswith("   ◆"):
            cur_zone = {"name": text.replace("◆", "").strip(), "hqs": []}
            zones.append(cur_zone)
            cur_hq = None
            continue
        if text.startswith("▶"):
            cur_hq = {"name": text.replace("▶", "").strip(), "people": []}
            if cur_zone is None:
                cur_zone = {"name": "Other", "hqs": []}
                zones.append(cur_zone)
            cur_zone["hqs"].append(cur_hq)
            continue
        # numbered person row
        if isinstance(c1, (int, float)):
            person = {
                "num": num(c1),
                "name": clean(row[2]),
                "role": clean(row[3]),
                "baseHQ": clean(row[4]),
                "reportsTo": clean(row[5]),
                "devices": num(row[6]) if row[6] not in (None, "") else None,
                "celluma": num(row[7]) if row[7] not in (None, "") else None,
                "notes": clean(row[8]),
            }
            if cur_hq is None:
                cur_hq = {"name": cur_zone["name"] if cur_zone else "Team", "people": []}
                (cur_zone or {"hqs": zones}).setdefault("hqs", []).append(cur_hq)
            cur_hq["people"].append(person)
    return zones


def parse_roster(ws):
    r = rows_of(ws)
    people = []
    for row in r[7:26]:  # rows 8..26
        n = row[1]
        if not isinstance(n, (int, float)):
            continue
        people.append({
            "num": num(n),
            "name": clean(row[2]),
            "designation": clean(row[3]),
            "baseHQ": clean(row[4]),
            "reportsTo": clean(row[5]),
            "zone": clean(row[6]),
            "notes": clean(row[11]) if len(row) > 11 else None,
        })
    summary = {}
    for row in r[26:30]:
        label = clean(row[1])
        val = None
        for c in row[2:]:
            if isinstance(c, (int, float)):
                val = num(c)
                break
        if label:
            summary[label.replace("•", "").strip()] = val
    return {"people": people, "summary": summary}


def parse_device_incentive(ws):
    r = rows_of(ws)

    def grab(start):
        out = []
        for row in r[start:start + 12]:
            name = clean(row[1])
            if not name or clean(row[2]) is None and not isinstance(row[2], (int, float)):
                continue
            out.append({
                "device": name,
                "quotation": num(row[2]),
                "standard": num(row[3]),
                "minimum": num(row[4]),
                "stdIncentive": num(row[5]),
                "minIncentive": num(row[6]),
                "aboveStd": clean(row[7]),
            })
        return out

    return {"salesperson": grab(7), "manager": grab(23)}


def parse_celluma_incentive(ws):
    r = rows_of(ws)
    out = []
    for row in r[7:20]:  # rows 8..20
        name = clean(row[1])
        if not name:
            continue
        out.append({
            "model": name,
            "sellingPrice": num(row[2]),
            "salespersonIncentive": num(row[3]),
            "managerIncentive": num(row[4]),
        })
    return out


def parse_esthemax_incentive(ws):
    r = rows_of(ws)

    def tiers(start):
        out = []
        for row in r[start:start + 6]:
            tier = clean(row[1])
            if not tier:
                continue
            out.append({
                "tier": tier,
                "min": num(row[2]),
                "max": clean(row[3]),
                "incentive": num(row[4]),
                "label": clean(row[5]),
            })
        return out

    return {"salesperson": tiers(7), "manager": tiers(20)}


def parse_device_cost(ws):
    r = rows_of(ws)
    out = []
    for row in r[5:18]:
        name = clean(row[1])
        if not name:
            continue
        out.append({
            "device": name,
            "landingCost": num(row[2]),
            "quotation": num(row[3]),
            "standard": num(row[4]),
            "minimum": num(row[5]),
        })
    return out


def parse_celluma_cost(ws):
    r = rows_of(ws)
    out = []
    for row in r[6:19]:
        name = clean(row[1])
        if not name:
            continue
        out.append({
            "model": name,
            "quotation": num(row[2]),
            "selling": num(row[3]),
        })
    return out


def parse_esthemax_cost(ws):
    r = rows_of(ws)
    sections = {"hydrojelly": [], "retail": [], "footMask": []}
    cur = None
    for row in r[4:47]:
        c1 = clean(row[1])
        if c1 is None:
            continue
        text = str(c1)
        if "HYDROJELLY MASK" in text:
            cur = "hydrojelly"; continue
        if "RETAIL HYDROJELLY" in text:
            cur = "retail"; continue
        if "COLLAGEN FOOT MASK" in text or (text == "Collagen Foot Mask"):
            cur = "footMask"
        if cur is None:
            continue
        if not isinstance(row[2], str):
            # variant rows have a pack string in col 2
            pass
        if clean(row[2]) is None:
            continue
        sections[cur].append({
            "variant": clean(row[1]),
            "pack": clean(row[2]),
            "landingCost": num(row[3]),
            "standardTotal": num(row[4]),
            "mrp": num(row[5]),
            "newMrp": num(row[6]),
            "minEXW": num(row[7]),
        })
    return sections


def parse_tc(ws):
    r = rows_of(ws)
    sections = []
    cur = None
    for row in r[5:]:
        c1 = clean(row[1])
        c2 = clean(row[2])
        if c1 is None:
            continue
        text = str(c1)
        if text[:2].strip().isdigit() and "." in text[:4]:
            cur = {"title": text.strip(), "items": []}
            sections.append(cur)
            continue
        if cur is not None and c2:
            cur["items"].append({"term": text.strip(), "detail": str(c2).strip()})
    return sections


def parse_comments(ws):
    r = rows_of(ws)
    out = []
    for row in r[6:28]:
        n = row[1]
        if not isinstance(n, (int, float)):
            continue
        out.append({
            "num": num(n),
            "raisedBy": clean(row[2]),
            "location": clean(row[3]),
            "topic": clean(row[4]),
            "comment": clean(row[5]),
            "resolution": clean(row[6]),
            "status": clean(row[7]),
        })
    return out


def parse_hq(ws):
    """Generic-ish parser for the regional HQ target sheets."""
    r = rows_of(ws)
    title = clean(r[1][1]) if len(r) > 1 else ws.title
    subtitle = clean(r[3][1]) if len(r) > 3 else None

    # Top summary block: header row is the one containing 'FY25-26 Actual'.
    # Values are on the next row, descriptive notes on the row after that.
    summary = []
    for i, row in enumerate(r[:12]):
        cells = [clean(c) for c in row]
        if any(isinstance(c, str) and "FY25-26 Actual" in c for c in cells):
            head = r[i]
            vals = r[i + 1] if i + 1 < len(r) else []
            notes = r[i + 2] if i + 2 < len(r) else []
            for j, h in enumerate(head):
                if not clean(h):
                    continue
                v = vals[j] if j < len(vals) else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                summary.append({
                    "label": clean(h),
                    "value": num(v) if isinstance(v, (int, float)) else clean(v),
                    "note": clean(notes[j]) if j < len(notes) else None,
                })
            break

    # Product plan tables — some HQs (e.g. West) carry several sub-plans, one
    # per rep. Capture every 'Product' header block, labelling each with the
    # nearest preceding section header (row starting with ▶ or 'ASM Plan —').
    plans = []
    for i, row in enumerate(r):
        cells = [clean(c) for c in row]
        if "Product" not in cells or not ("FY26-27" in cells or "FY25-26" in cells):
            continue
        # find a label by scanning upward for a section header
        label = None
        for back in range(i - 1, max(i - 6, -1), -1):
            t = clean(r[back][1])
            if isinstance(t, str) and t.strip():
                st = t.strip()
                if st.startswith("▶") or "Plan" in st or "ASM" in st or "Role" in st:
                    label = st.replace("▶", "").strip()
                    break
        rows = []
        for prow in r[i + 1:i + 30]:
            name = clean(prow[1])
            if not name:
                continue
            if str(name).upper().startswith("TOTAL"):
                rows.append({
                    "product": "TOTAL",
                    "fy2526": num(prow[2]) if isinstance(prow[2], (int, float)) else None,
                    "fy2627": num(prow[3]) if isinstance(prow[3], (int, float)) else None,
                    "totalValue": num(prow[5]) if isinstance(prow[5], (int, float)) else None,
                    "isTotal": True,
                })
                break
            if not any(isinstance(prow[k], (int, float)) for k in (2, 3, 4, 5)):
                break
            rows.append({
                "product": name,
                "fy2526": num(prow[2]) if isinstance(prow[2], (int, float)) else prow[2],
                "fy2627": num(prow[3]) if isinstance(prow[3], (int, float)) else prow[3],
                "deviceValue": num(prow[4]) if isinstance(prow[4], (int, float)) else None,
                "totalValue": num(prow[5]) if isinstance(prow[5], (int, float)) else None,
                "notes": clean(prow[9]) if len(prow) > 9 else None,
            })
        if rows:
            plans.append({"label": label, "rows": rows})
    # Backwards-compatible flat plan = first table's rows.
    plan = plans[0]["rows"] if plans else []

    # Quarterly target table (Units / Value rows)
    quarterly = []
    for i, row in enumerate(r):
        cells = [clean(c) for c in row]
        if "Basis" in cells and "Annual" in cells and "Q1" in cells:
            for qrow in r[i + 1:i + 5]:
                basis = clean(qrow[1])
                # Only real quarterly rows: numeric annual + numeric quarters.
                if not basis or not isinstance(qrow[2], (int, float)):
                    break
                quarterly.append({
                    "basis": basis,
                    "annual": num(qrow[2]),
                    "q1": num(qrow[3]),
                    "q2": num(qrow[4]),
                    "q3": num(qrow[5]),
                    "q4": num(qrow[6]),
                })
            break

    # Plan highlights (emoji bullet lines)
    highlights = []
    for row in r:
        c1 = clean(row[1])
        if isinstance(c1, str) and c1 and c1[0] in "📊💎💰🆕🎯👥🗺🔑⭐✅":
            highlights.append(c1)

    return {
        "sheet": ws.title,
        "title": title,
        "subtitle": subtitle,
        "summary": summary,
        "plan": plan,
        "plans": plans,
        "quarterly": quarterly,
        "highlights": highlights,
    }


# --------------------------------------------------------------------------
# Esthemax price workbook (Salon / Doctor)
# --------------------------------------------------------------------------
def parse_price_market(ws):
    """Positional parse — the price sheets have duplicate column names
    (Old Structure vs New Structure both have 'MRP', 'Effective Net Price'),
    so we keep values as an array aligned to the header, never a name-keyed
    dict (which would silently drop columns)."""
    r = rows_of(ws)
    band = [clean(c) if c is not None else "" for c in r[0]]   # row 1: Old/New band
    header = [clean(c) if c is not None else "" for c in r[1]]  # row 2: column names
    ncol = len(header)
    groups = {"HYDROJELLYMASK": [], "RETAIL HYDROJELLYMASK": [], "Foot Mask": []}
    cur = None
    for row in r[2:]:
        first = clean(row[0])
        if isinstance(first, str):
            key = first.upper().replace(" ", "")
            if key == "HYDROJELLYMASK":
                cur = "HYDROJELLYMASK"; continue
            if key == "RETAILHYDROJELLYMASK":
                cur = "RETAIL HYDROJELLYMASK"; continue
            if key == "FOOTMASK":
                cur = "Foot Mask"; continue
        if cur is None or not isinstance(first, (int, float)):
            continue
        values = []
        for idx in range(ncol):
            v = row[idx] if idx < len(row) else None
            values.append(num(v) if isinstance(v, (int, float)) else clean(v))
        groups[cur].append({"srNo": num(first), "name": clean(row[1]), "values": values})
    return {"band": band, "columns": header, "groups": groups}


def declean(v):
    """Clean a Demo-Machine cell: format dates, and undo the source's
    find/replace corruption where the letter 'v'/'V' became 'Booked'
    (e.g. 'Bookedossman' -> 'Vossman'). The literal status 'Booked' is kept."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d %b %Y")
    if isinstance(v, str):
        s = v.strip()
        if s and s != "Booked":
            s = s.replace("Booked", "V")
        return s
    return num(v) if isinstance(v, (int, float)) else v


def parse_demo_table(ws, col_idxs, headers, name_col, data_start=1):
    data = rows_of(ws)
    out = []
    for row in data[data_start:]:
        first = row[name_col] if name_col < len(row) else None
        if first is None or (isinstance(first, str) and not first.strip()):
            continue
        out.append([declean(row[i]) if i < len(row) else None for i in col_idxs])
    return {"columns": headers, "rows": out}


def parse_dropdowns(wb):
    """Read the 'Drop Down' sheet — column A = Status list, column B = Condition
    list (header in row 1)."""
    ws = wb["Drop Down"]
    rows = rows_of(ws)
    status, condition = [], []
    for row in rows[1:]:
        a = declean(row[0]) if len(row) > 0 else None
        b = declean(row[1]) if len(row) > 1 else None
        if isinstance(a, str) and a.strip():
            status.append(a.strip())
        if isinstance(b, str) and b.strip():
            condition.append(b.strip())
    return {"status": status, "condition": condition}


def parse_refs(wb):
    """Master lookup lists from the 'DATA Sheet':
    A = States/UTs (Location), B = Employees (salesperson/assigned-to),
    C = Devices, D = Serial numbers (C & D are row-aligned)."""
    ws = wb["DATA Sheet"]
    rows = rows_of(ws)[1:]  # skip header row

    def col(idx):
        seen, out = set(), []
        for row in rows:
            v = declean(row[idx]) if idx < len(row) else None
            if isinstance(v, str) and v.strip() and v not in seen:
                seen.add(v); out.append(v)
        return out

    device_serials = []
    for row in rows:
        dev = declean(row[2]) if len(row) > 2 else None
        ser = declean(row[3]) if len(row) > 3 else None
        if isinstance(dev, str) and dev.strip():
            device_serials.append({"device": dev, "serial": ser if isinstance(ser, str) else ""})
    return {
        "states": col(0),
        "employees": col(1),
        "devices": col(2),
        "serials": col(3),
        "deviceSerials": device_serials,
    }


def parse_demo(wb):
    # Current booking status (header at row 3; the daily grid columns are skipped).
    current = parse_demo_table(
        wb["Current April-July 2026"],
        [0, 1, 3, 4, 5, 6],
        ["Device", "Serial No", "Status", "Location", "Remarks", "Manager"],
        name_col=0, data_start=3)
    # Machine working-condition details.
    status = parse_demo_table(
        wb["Demo Machine"],
        [0, 1, 2, 3, 4, 5, 6, 7, 8, 10],
        ["Machine", "Serial No", "Location", "Purpose", "Doctor", "Salesperson",
         "Confirmed by", "Status", "Updated", "Condition"],
        name_col=0)
    # Movement / dispatch log.
    movement = parse_demo_table(
        wb["Machine Movement Data "],
        [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13],
        ["Date", "Device", "Serial No", "From", "To", "Assigned To", "Purpose",
         "Expected Return", "Actual Return", "Demo Status", "Cond. Out",
         "Cond. Return", "Accessories"],
        name_col=2)
    # Flight-case / packing condition.
    packing = parse_demo_table(
        wb["Packing condition "],
        list(range(0, 16)),
        ["Device", "Serial No", "Flight Case ID", "Foam Condition", "Lock Condition",
         "Handle Condition", "Wheel Condition", "Exterior Condition", "Interior Condition",
         "Last Checked By", "Last Checked Date", "Status", "Remarks",
         "Flight Case On Receipt", "Missing Items", "Damage Reported"],
        name_col=0)
    return {"current": current, "status": status, "movement": movement, "packing": packing}


def parse_challan_refs(wb):
    """Address book for the Delivery Challan module:
    - consignors (From): the Primelaze offices in the 'Database' sheet
    - consignees (To): recipients in the 'DATA' sheet (name, address, declaration)
    """
    def tidy(v):
        if not isinstance(v, str):
            return ""
        return " ".join(v.replace("\xa0", " ").split()).strip()

    consignors = []
    for row in rows_of(wb["Database"])[1:]:
        name = tidy(row[1]) if len(row) > 1 else ""
        addr = tidy(row[2]) if len(row) > 2 else ""
        if name and "primelaze" in name.lower():
            consignors.append({"name": name, "addr": addr})

    consignees = []
    seen = set()
    for row in rows_of(wb["DATA"])[1:]:
        name = tidy(row[1]) if len(row) > 1 else ""
        addr = tidy(row[3]) if len(row) > 3 else ""
        purpose = tidy(row[5]) if len(row) > 5 else ""
        if name and name not in seen:
            seen.add(name)
            consignees.append({"name": name, "addr": addr, "purpose": purpose})

    return {"from": consignors, "to": consignees}


def parse_order(ws):
    """Esthemax procurement planner (Order Summary sheet).

    Required = round(avg*dermaMonths) + round(avg*salonMonths)
    Landing/unit = unitUSD*usdInr*(1+customsRate) + transport
    To buy = max(0, Required - currentStock);  Money = toBuy * landing
    We keep the 15 months of raw sales so the app can recompute the 6-mo
    average (and everything downstream) when the parameters are changed.
    """
    r = rows_of(ws)
    params = {"dermaMonths": 3, "salonMonths": 1.5, "usdInr": 96.99, "customsRate": 0.44}
    for row in r[:8]:
        label = clean(row[1])
        if label == "Derma months":
            params["dermaMonths"] = num(row[3]); params["usdInr"] = num(row[7])
        elif label == "Salon months":
            params["salonMonths"] = num(row[3]); params["customsRate"] = num(row[7])

    # header row containing 'Item'
    head_i = None
    for i, row in enumerate(r):
        if clean(row[1]) == "Item" and clean(row[2]) == "Category":
            head_i = i; break
    months = [clean(c) for c in r[head_i][3:18]] if head_i is not None else []

    items = []
    for row in r[(head_i + 1):]:
        name = clean(row[1])
        if not name:
            continue
        if str(name).upper().startswith("TOTAL"):
            break
        monthly = [num(row[j]) if isinstance(row[j], (int, float)) else None for j in range(3, 18)]
        val = lambda j, default=0: num(row[j]) if isinstance(row[j], (int, float)) else default
        # requiredStock is authoritative from the workbook: product masks follow
        # round(avg*derma)+round(avg*salon), but accessories carry a manually-set
        # target that isn't derivable from the average — so we store it as-is.
        items.append({
            "name": name,
            "category": clean(row[2]) or "Other",
            "monthly": monthly,
            "sixMoAvg": val(18),
            "requiredStock": val(21),
            "currentStock": val(22),
            "toBuyRef": val(23),
            "unitUSD": val(24),
            "transport": val(27),
            "moneyRef": val(29),
        })
    return {"params": params, "months": months, "items": items}


def apply_amendments(data):
    """Post-extraction corrections for events that happened after the source
    workbook was authored. Applied here (not by editing the .xlsx) because an
    openpyxl round-trip would strip the workbook's 131 cached formula values.
    Deterministic: extraction always starts from the pristine workbook, so each
    run reproduces the same amended result.

    -- Kuldeep Singh (North-1 ZSM, Chandigarh) resigned. --
    Seat marked vacant (matches how the workbook already shows vacancies);
    resignation recorded in notes; FY26-27 targets held on the vacant seat
    until backfill; Akshay Jain re-points to the vacant ZSM; headcounts move
    active 12->11, vacant 4->5.
    """
    RESIGNED = "Kuldeep Singh"
    VACANT = "Vacant North-1 ZSM"
    NOTE = "Kuldeep Singh resigned — backfill pending (HIGH PRIORITY)."

    def swap_k(s):
        # Ordered longest-first replacements in a single pass. Substituted text
        # uses a sentinel so later rules can't re-hit it (chained str.replace
        # would otherwise mangle "(Kuldeep resigned)").
        if not isinstance(s, str) or "Kuldeep" not in s:
            return s
        reps = [
            (RESIGNED + " (ZSM, Chandigarh)", "Vacant ZSM «KR»"),
            (RESIGNED + " (ZSM)", "Vacant ZSM «KR»"),
            ("Kuldeep is pure ZSM", "the seat is now vacant «KR»"),
            ("Kuldeep is the Zonal Sales Manager", "This ZSM seat is vacant «KR»"),
            ("Kuldeep ZSM", "the vacant ZSM"),
            ("Kuldeep takes", "the vacant ZSM covers"),
            (RESIGNED, "Vacant ZSM «KR»"),
            ("Kuldeep", "the vacant ZSM"),
        ]
        for a, b in reps:
            s = s.replace(a, b)
        return s.replace("«KR»", "(Kuldeep resigned)")

    # KPIs
    k = data["kpis"]
    if isinstance(k.get("activeReps"), (int, float)):
        k["activeReps"] = k["activeReps"] - 1
    k["vacant"] = (k.get("vacant") or 0) + 1
    vn = k.get("vacantNote") or ""
    k["vacantNote"] = (vn + " + North-1 ZSM (Chandigarh)") if vn else "North-1 ZSM (Chandigarh)"

    # Team roster
    for p in data["roster"]["people"]:
        if p.get("name") == RESIGNED:
            p["name"] = VACANT
            p["notes"] = NOTE + " Was Punjab IC: Chandigarh, HP, J&K, Uttarakhand."
        if p.get("reportsTo") == RESIGNED:
            p["reportsTo"] = VACANT
    summ = data["roster"].get("summary", {})
    for key, val in list(summ.items()):
        if not isinstance(val, (int, float)):
            continue
        if "Active" in key:
            summ[key] = val - 1
        elif "Vacant" in key:
            summ[key] = val + 1

    # Master Dashboard zones (HQ name + person rows)
    for z in data["zones"]:
        for hq in z["hqs"]:
            hq["name"] = swap_k(hq.get("name"))
            for p in hq["people"]:
                if p.get("name") == RESIGNED:
                    p["name"] = VACANT
                    p["notes"] = NOTE + " Targets held on the vacant seat until backfilled. " + (p.get("notes") or "")
                if p.get("reportsTo") == RESIGNED:
                    p["reportsTo"] = VACANT

    # Punjab HQ target sheet (descriptive text)
    for h in data["hqTargets"]:
        if "Punjab" in (h.get("sheet") or ""):
            h["subtitle"] = swap_k(h.get("subtitle"))
            h["highlights"] = [swap_k(x) for x in h.get("highlights", [])]
            for pl in h.get("plans", []):
                pl["label"] = swap_k(pl.get("label"))

    # -- Divisions: Lubdha & Puneet in Salon/Spa, everyone else Derma. --
    SALON = {"Ms. Lubdha", "Puneet"}
    for p in data["roster"]["people"]:
        p["division"] = "Salon/Spa" if p.get("name") in SALON else "Derma"
    # Add Puneet as a new Salon/Spa person (not present in the source workbook).
    next_roster = max((p.get("num") or 0 for p in data["roster"]["people"]), default=0) + 1
    data["roster"]["people"].append({
        "num": next_roster,
        "name": "Puneet",
        "designation": "Sales – Salon/Spa",
        "baseHQ": "—",
        "reportsTo": "Arjun",
        "zone": "Salon/Spa Division",
        "division": "Salon/Spa",
        "notes": "Salon/Spa division. Details pending (base HQ / territory).",
    })
    for key, val in list(summ.items()):
        if not isinstance(val, (int, float)):
            continue
        if "Active" in key or "Total" in key:
            summ[key] = val + 1  # Puneet joins as active

    # Comments log audit entry
    next_num = max((c.get("num") or 0 for c in data["comments"]), default=0) + 1
    data["comments"].append({
        "num": next_num,
        "raisedBy": "HR",
        "location": "Team Roster · C8",
        "topic": "North-1 ZSM resignation",
        "comment": "Kuldeep Singh resigned.",
        "resolution": ("Seat marked vacant (North-1 ZSM, Chandigarh). Active 12→11, vacant 4→5. "
                       "FY26-27 targets (30 devices / 20 Celluma) held on the vacant seat; Akshay "
                       "Jain now reports to the vacant ZSM until backfill."),
        "status": "✅ Done",
    })
    return data


def main():
    dash = load_workbook(DASHBOARD, data_only=True)
    prices = load_workbook(PRICES, data_only=True)
    order_wb = load_workbook(ORDER, data_only=True)
    demo_wb = load_workbook(DEMO, data_only=True)
    challan_wb = load_workbook(CHALLAN, data_only=True)

    hq_sheets = [s for s in dash.sheetnames if s.endswith("HQ")]

    data = {
        "meta": {
            "company": "Primelaze",
            "fiscalYear": "FY 2026-27",
            "title": "Primelaze Unified Dashboard",
            "source": [
                "Primelaze_Unified_Dashboard_FY2627.xlsx",
                "Esthemax_prices_working.xlsx",
                "Esthemax_Order_Calculation.xlsx",
                "Demo_Machine_Booking_Status.xlsx",
            ],
        },
        "kpis": parse_kpis(dash["Master Dashboard"]),
        "zones": parse_master_targets(dash["Master Dashboard"]),
        "roster": parse_roster(dash["Team Roster"]),
        "incentives": {
            "device": parse_device_incentive(dash["Device Incentive"]),
            "celluma": parse_celluma_incentive(dash["Celluma Incentive"]),
            "esthemax": parse_esthemax_incentive(dash["Esthemax Incentive"]),
            "terms": parse_tc(dash["Incentive T&C"]),
        },
        "costs": {
            "device": parse_device_cost(dash["Device Cost"]),
            "celluma": parse_celluma_cost(dash["Celluma Cost"]),
            "esthemax": parse_esthemax_cost(dash["Esthemax Cost"]),
        },
        "hqTargets": [parse_hq(dash[s]) for s in hq_sheets],
        "comments": parse_comments(dash["Comments Log"]),
        "esthemaxPrices": {
            "salon": parse_price_market(prices["Salon Market"]),
            "doctor": parse_price_market(prices["Doctor Market"]),
        },
        "esthemaxOrder": parse_order(order_wb["Order Summary"]),
        "demoMachines": parse_demo(demo_wb),
        "dropdowns": parse_dropdowns(demo_wb),
        "refs": parse_refs(demo_wb),
        "challanRefs": parse_challan_refs(challan_wb),
    }

    apply_amendments(data)

    out_json = os.path.join(ROOT, "src", "data", "data.json")
    os.makedirs(os.path.dirname(out_json), exist_ok=True)
    with open(out_json, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print("Wrote:")
    print(" ", out_json)
    print("Next: run `node scripts/encrypt_data.js` to refresh the encrypted")
    print("      assets/js/data.enc.js payload the website loads.")
    print(f"KPIs: {data['kpis']}")
    print(f"Zones: {len(data['zones'])}  HQ sheets: {len(data['hqTargets'])}")
    print(f"Roster: {len(data['roster']['people'])} people")
    print(f"Device incentive SP rows: {len(data['incentives']['device']['salesperson'])}")
    print(f"Esthemax cost hydrojelly: {len(data['costs']['esthemax']['hydrojelly'])}")
    print(f"Order planner items: {len(data['esthemaxOrder']['items'])}  params: {data['esthemaxOrder']['params']}")


if __name__ == "__main__":
    main()
